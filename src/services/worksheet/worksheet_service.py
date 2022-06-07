import logging
import math

from openpyxl import Workbook

from openpyxl.styles import Border, Font, Alignment, GradientFill
from openpyxl.styles.borders import Side
from openpyxl.utils import get_column_letter


class WorksheetService:
    COLUMNS_MARGIN = 1
    MARGIN_COLUMN_WIDTH = 5
    MIN_COLUMN_LENGTH = 10
    ROWS_MARGIN = 1

    workbook = None
    worksheet = None
    headers: list = []
    rows: list = []
    title: str = ""
    max_columns: int = 0

    def __init__(self, headers: list, rows: list, title: str = ""):
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Dados"
        self.headers = headers
        self.rows = rows
        self.title = title
        self.refresh()

    def refresh(self):
        data_starting_row = 1 + self.ROWS_MARGIN
        self.max_columns = self.get_max_columns()

        if self.title != "":
            self.set_title(text=self.title, starting_row=data_starting_row)
            data_starting_row += 1

        if len(self.headers) > 0:
            self.set_headers(headers=self.headers, starting_row=data_starting_row)
            data_starting_row += 1

        self.set_rows(rows=self.rows, starting_row=data_starting_row)
        self.adjust_column_filters(data_starting_row=data_starting_row-1)

    def adjust_column_filters(self, data_starting_row: int = 1):
        ending_row = data_starting_row + len(self.rows)
        starting_column = get_column_letter(1 + self.COLUMNS_MARGIN)
        ending_column = get_column_letter(self.COLUMNS_MARGIN + self.max_columns)

        reference = f"{starting_column}{data_starting_row}:{ending_column}{ending_row}"
        logging.debug(f"Data cells reference: {reference}")

        self.worksheet.auto_filter.ref = f"{starting_column}{data_starting_row}:{ending_column}{ending_row}"

        for column_index, column_cells in enumerate(self.worksheet.columns):
            if column_index < self.COLUMNS_MARGIN:
                continue

            if column_index > self.COLUMNS_MARGIN + len(self.headers):
                continue

            unique_column_values = list(set(column_cells[data_starting_row:]))
            self.worksheet.auto_filter.add_filter_column(column_index, vals=unique_column_values)

    def save(self, filepath):
        self.adjust_columns_width()
        self.workbook.save(filename=filepath)

    @staticmethod
    def get_string_computed_length(input_string: str, font_size: int) -> int:
        size = 0.0

        upper_chars_count = sum(map(str.isupper, input_string))
        logging.debug(f"total upper chars in '{input_string}': {upper_chars_count}")
        size += upper_chars_count * 1.18

        lower_chars_count = sum(map(str.islower, input_string))
        logging.debug(f"total lower chars in '{input_string}': {lower_chars_count}")
        size += lower_chars_count * 0.85

        spaces_count = input_string.count(" ")
        logging.debug(f"total space chars in '{input_string}': {spaces_count}")
        size += spaces_count * 1.33

        chars_left = len(input_string) - upper_chars_count - lower_chars_count - spaces_count
        logging.debug(f"total chars left in '{input_string}': {chars_left}")
        size += chars_left * 1.01

        base_font_size = 11.0
        font_size_factor = ((float(font_size) / base_font_size) - 1.0) * 10.0

        return math.ceil(size + font_size_factor)

    def adjust_columns_width(self):
        column_start_index = (0 if self.title == "" else 1) + self.COLUMNS_MARGIN
        for column_index, column_cells in enumerate(self.worksheet.columns):
            column_letter = get_column_letter(column_cells[0].column)

            if column_index == 0:
                self.worksheet.column_dimensions[column_letter].width = self.MARGIN_COLUMN_WIDTH

                continue

            row_lengths = [
                self.get_string_computed_length(str(cell.value), cell.font.size) for cell in column_cells[column_start_index:]
            ]
            logging.debug(f"row lengths for column #{column_index}: {row_lengths}")
            computed_column_length = max([
                self.get_string_computed_length(str(cell.value), cell.font.size) for cell in column_cells[column_start_index:]
            ])
            new_column_length = max(self.MIN_COLUMN_LENGTH, computed_column_length)

            self.worksheet.column_dimensions[column_letter].width = new_column_length * 1.23

    def get_max_columns(self):
        sizes = [len(row) for row in self.rows]
        sizes.append(len(self.headers))

        return max(sizes)

    def get_cell(self, row: int, column: int):
        return self.worksheet.cell(row=row, column=column)

    def set_cell_value(self, row: int, column: int, value):
        return self.worksheet.cell(row=row, column=column, value=value)

    def set_title(self, text: str, starting_row: int = 1):
        starting_column = 1 + self.COLUMNS_MARGIN
        for column_index in range(starting_column, self.max_columns + starting_column):
            title_cell = self.get_cell(row=starting_row, column=column_index)
            title_cell.border = self.get_border(
                top=True,
                left=column_index == starting_column,
                right=column_index == self.max_columns + starting_column - 1,
                bottom=True
            )

        if self.max_columns > 1:
            self.worksheet.merge_cells(
                start_row=starting_row,
                start_column=starting_column,
                end_row=starting_row,
                end_column=starting_column + self.max_columns - 1
            )

        title_cell = self.set_cell_value(row=starting_row, column=starting_column, value=text)
        title_cell.font = self.get_font(size=14, bold=True, color="EAEAEA")
        title_cell.alignment = self.get_alignment(horizontal="center")
        title_cell.fill = self.get_fill(start_color="0283B6", end_color="0AA3E5")

    def set_headers(self, headers: list, starting_row: int = 1):
        starting_column = 1 + self.COLUMNS_MARGIN
        for index, header in enumerate(headers):
            header_cell = self.set_cell_value(row=starting_row, column=(index+starting_column), value=header)
            header_cell.font = self.get_font(size=12, bold=True)
            header_cell.border = self.get_border(top=True, left=True, right=True, bottom=True)
            header_cell.alignment = self.get_alignment(horizontal="center")

    def set_rows(self, rows: list, starting_row: int = 1):
        last_row_pointer = starting_row + len(rows) - 1
        starting_column = 1 + self.COLUMNS_MARGIN

        for row_index, row in enumerate(rows):
            row_pointer = starting_row + row_index

            for column_index, value in enumerate(row):
                cell = self.set_cell_value(row=row_pointer, column=column_index+starting_column, value=value)
                cell.border = self.get_border(
                    left=True,
                    right=True,
                    bottom=row_pointer == last_row_pointer
                )

    @staticmethod
    def get_font(size: int = 11, bold: bool = False, italic: bool = False, color: str = 'FF000000'):
        return Font(
            name='Calibri',
            size=size,
            bold=bold,
            italic=italic,
            vertAlign=None,
            underline='none',
            strike=False,
            color=color
        )

    @staticmethod
    def get_fill(start_color: str = 'FFFFFFFF', end_color: str = 'FF000000'):
        return GradientFill(
            type="linear",
            degree="45",
            stop=[start_color, end_color],
        )

    @staticmethod
    def get_alignment(horizontal: str = "left", vertical: str = "top", wrap: bool = False):
        return Alignment(
            horizontal=horizontal,
            vertical=vertical,
            text_rotation=0,
            wrap_text=wrap,
            shrink_to_fit=False,
            indent=0
        )

    @staticmethod
    def get_border_side(color: str = 'FF000000', border_style: str = 'thin') -> Side:
        return Side(
            border_style=border_style,
            color=color
        )

    def get_border(
        self,
        color: str = 'FF000000',
        border_style: str = 'thin',
        left: bool = False,
        right: bool = False,
        top: bool = False,
        bottom: bool = False
    ):
        side = self.get_border_side(
            border_style=border_style,
            color=color
        )

        border = Border(
            top=side if top else None,
            left=side if left else None,
            right=side if right else None,
            bottom=side if bottom else None,
        )

        return border
